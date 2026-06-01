"""
Excel workbook polish: Tables (AutoFilter), dropdown validations, freeze panes, column widths.

Used by build_tabular_dataset, build_canonical_taxonomy, and build_analytical_master.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from src.enrichment.academic_taxonomy import taxonomy_validation_lists
from src.enrichment.canonical_taxonomy import (
    CANONICAL_CATEGORIES,
    IGNORE_LABEL,
    META_SYNTHESIS_CANONICAL_METHOD,
    METHOD_CANONICAL_MAP,
    UNCATEGORIZED,
)

FILTER_LISTS_SHEET = "_FilterLists"
MAX_COL_WIDTH = 52
MIN_COL_WIDTH = 10

PHASE0_SHEET_NAMES: tuple[str, ...] = (
    "5_Verification_Claims",
    "6_Verification_Results",
    "7_Prediction_Inventory",
)

# Literature first, Phase 0 verification/inventory last (before hidden _FilterLists).
CLEAN_WORKBOOK_SHEET_ORDER: tuple[str, ...] = (
    "0_Dashboard_Analysis_Control",
    "1_Articles_Master",
    "2_Main_Findings",
    "3_Confounders",
    "Canonical_View",
    *PHASE0_SHEET_NAMES,
    FILTER_LISTS_SHEET,
)

CANONICAL_RECORD_TYPES: tuple[str, ...] = ("finding", "confounder")
CANONICAL_EFFECT_TRENDS: tuple[str, ...] = ("Positive", "Negative", "Null", IGNORE_LABEL)
CANONICAL_METADATA_FLAGS: tuple[str, ...] = (
    "empirical_finding",
    "theoretical_meta_synthesis",
)

POLICY_IMPACT_DIRECTIONS: tuple[str, ...] = (
    "Improving (positive predictor)",
    "Corrective (negative association)",
    "Mixed / Context-dependent",
    "Descriptive / Null",
    "Not Reported",
    "[Not Explicitly Stated]",
)

# ── Phase 0 verification / prediction sheet enums ──
NORMALIZED_ISA_OPTIONS: tuple[str, ...] = (
    "PISA",
    "TIMSS",
    "PIRLS",
    "ICCS",
    "ICILS",
    "PIAAC",
    "TALIS",
)

ANALYSIS_LEVEL_OPTIONS: tuple[str, ...] = (
    "student",
    "school",
    "country",
)

PV_RULE_OPTIONS: tuple[str, ...] = (
    "rubin_rules",
    "single_pv",
    "average_pv",
    "all_pv_separate",
    "not_applicable",
    "not_reported",
)

CATALOG_MATCH_CONFIDENCE_OPTIONS: tuple[str, ...] = ("High", "Medium", "Low")

VERIFICATION_ELIGIBLE_OPTIONS: tuple[str, ...] = ("TRUE", "FALSE")

VERIFICATION_PRIORITY_OPTIONS: tuple[str, ...] = (
    "P1_pilot",
    "P2_core",
    "P3_extended",
    "P4_deferred",
)

CHECK_TYPE_OPTIONS: tuple[str, ...] = (
    "descriptive",
    "weighted_mean",
    "correlation_direction",
    "correlation_magnitude",
    "regression_coefficient_sign",
    "metric_replication",
    "sample_size",
)

EFFECT_DIRECTION_OPTIONS: tuple[str, ...] = (
    "Positive",
    "Negative",
    "Null",
    "Unknown",
)

PAPER_SIGNIFICANCE_OPTIONS: tuple[str, ...] = (
    "Significant",
    "Not_Significant",
    "Not_Reported",
)

MATCH_STATUS_OPTIONS: tuple[str, ...] = (
    "Full_Match",
    "Partial_Match",
    "Direction_Mismatch",
    "Discrepancy",
    "Flagged_Anomaly",
    "Not_Runnable",
    "Inconclusive",
)

FEATURE_STORE_STATUS_OPTIONS: tuple[str, ...] = (
    "Not_Started",
    "Planned",
    "Partial",
    "Complete",
)

PIPELINE_PRIORITY_OPTIONS: tuple[str, ...] = (
    "Critical",
    "High",
    "Medium",
    "Low",
    "Deferred",
)

SHEET5_COLUMNS: tuple[str, ...] = (
    "claim_id",
    "file_name",
    "doi",
    "finding_row_hash",
    "claim_text",
    "normalized_isa",
    "normalized_cycle",
    "normalized_grade",
    "normalized_domain",
    "countries_iso3",
    "analysis_level",
    "literature_outcome_label",
    "literature_predictor_labels",
    "official_outcome_vars",
    "official_predictor_vars",
    "weight_var",
    "weight_profile_key",
    "pv_rule",
    "pv_draw_used_by_paper",
    "catalog_file_id",
    "catalog_match_confidence",
    "verification_eligible",
    "verification_priority",
    "bridge_provenance",
    "created_at",
)

SHEET6_COLUMNS: tuple[str, ...] = (
    "result_id",
    "claim_id",
    "check_type",
    "check_subtype",
    "paper_reported_value",
    "paper_reported_direction",
    "paper_significance",
    "replicated_value",
    "replicated_se",
    "replicated_p_value",
    "replicated_direction",
    "match_delta",
    "match_status",
    "alignment_narrative",
    "n_analytic",
    "pv_rule_applied",
    "weight_var_applied",
    "country_filter_applied",
    "evidence_path",
    "evidence_artifact_list",
    "engine_version",
    "run_timestamp",
    "run_duration_sec",
    "failure_code",
)

SHEET7_COLUMNS: tuple[str, ...] = (
    "inventory_id",
    "target_domain",
    "canonical_variable",
    "registry_outcome_key",
    "isa_coverage",
    "cycle_coverage",
    "predictor_category_mix",
    "top_predictors_canonical",
    "n_supporting_studies",
    "n_verified_claims",
    "dominant_algorithms_in_literature",
    "dominant_canonical_method",
    "effect_direction_consensus",
    "microdata_ready",
    "feature_store_status",
    "multi_output_pipeline_priority",
    "recommended_model_class",
    "measurement_invariance_note",
    "last_inventory_update",
)

POLICY_DOMAINS: tuple[str, ...] = (
    "Curriculum & Instruction",
    "Teacher Policy & Workforce",
    "School Finance & Resources",
    "Socio-Economic Status & Equity",
    "Digital Learning & ICT",
    "Language & Immigrant Policy",
    "Family Engagement",
    "Student Wellbeing & Climate",
    "Motivation & Self-Efficacy",
    "Gender & Demographic Equity",
    "System Governance",
    "Peer Effects & Tracking",
    "Retention & Pathways",
    "Civic Education",
    "[Not Explicitly Stated]",
)


def canonical_excel_validation_lists() -> dict[str, tuple[str, ...]]:
    """Dropdown vocabularies for Canonical_View and related sheets."""
    methods: list[str] = sorted(METHOD_CANONICAL_MAP.keys())
    methods.extend([META_SYNTHESIS_CANONICAL_METHOD, IGNORE_LABEL, UNCATEGORIZED])
    variables = sorted(set(CANONICAL_CATEGORIES.keys()) | {UNCATEGORIZED, IGNORE_LABEL})
    lists = taxonomy_validation_lists()
    predictor_cats = lists.get("predictor_category", ())
    return {
        "record_type": CANONICAL_RECORD_TYPES,
        "study_filter_type": lists["study_filter_type"],
        "canonical_method": tuple(dict.fromkeys(methods)),
        "canonical_variable": tuple(variables),
        "metadata_filter_flag": CANONICAL_METADATA_FLAGS,
        "effect_trend": CANONICAL_EFFECT_TRENDS,
        "target_domain": lists["target_domain"],
        "predictor_category": predictor_cats,
        "ml_family": lists["ml_family"],
        "target_dimension": lists["target_dimension"],
        "predictor_level": lists["predictor_level"],
        "pv_filter_label": lists["pv_filter_label"],
        "md_filter_label": lists["md_filter_label"],
        "weights_filter": lists["weights_filter"],
        "Policy_Domain": POLICY_DOMAINS,
        "Policy_Impact_Direction": POLICY_IMPACT_DIRECTIONS,
        "Study_Filter_Type": lists["study_filter_type"],
        "ML_Family": lists["ml_family"],
        "Target_Domain": lists["target_domain"],
        "Target_Dimension": lists["target_dimension"],
    }


def phase0_excel_validation_lists() -> dict[str, tuple[str, ...]]:
    """Dropdown vocabularies for Sheets 5–7 (Phase 0 verification layer)."""
    base = canonical_excel_validation_lists()
    return {
        **base,
        "normalized_isa": NORMALIZED_ISA_OPTIONS,
        "analysis_level": ANALYSIS_LEVEL_OPTIONS,
        "pv_rule": PV_RULE_OPTIONS,
        "catalog_match_confidence": CATALOG_MATCH_CONFIDENCE_OPTIONS,
        "verification_eligible": VERIFICATION_ELIGIBLE_OPTIONS,
        "verification_priority": VERIFICATION_PRIORITY_OPTIONS,
        "check_type": CHECK_TYPE_OPTIONS,
        "effect_direction": EFFECT_DIRECTION_OPTIONS,
        "paper_significance": PAPER_SIGNIFICANCE_OPTIONS,
        "match_status": MATCH_STATUS_OPTIONS,
        "feature_store_status": FEATURE_STORE_STATUS_OPTIONS,
        "pipeline_priority": PIPELINE_PRIORITY_OPTIONS,
    }


PHASE0_VERIFICATION_SHEET_CONFIG: dict[str, dict[str, object]] = {
    "5_Verification_Claims": {
        "table": "tbl_VerificationClaims",
        "dropdowns": (
            "normalized_isa",
            "normalized_domain",
            "analysis_level",
            "pv_rule",
            "catalog_match_confidence",
            "verification_eligible",
            "verification_priority",
        ),
    },
    "6_Verification_Results": {
        "table": "tbl_VerificationResults",
        "dropdowns": (
            "check_type",
            "paper_reported_direction",
            "paper_significance",
            "replicated_direction",
            "match_status",
            "pv_rule_applied",
        ),
    },
    "7_Prediction_Inventory": {
        "table": "tbl_PredictionInventory",
        "dropdowns": (
            "target_domain",
            "canonical_variable",
            "effect_direction_consensus",
            "microdata_ready",
            "feature_store_status",
            "multi_output_pipeline_priority",
        ),
    },
}


_DROPDOWN_LIST_ALIASES: dict[str, str] = {
    "normalized_domain": "target_domain",
    "paper_reported_direction": "effect_direction",
    "replicated_direction": "effect_direction",
    "pv_rule_applied": "pv_rule",
    "microdata_ready": "verification_eligible",
    "multi_output_pipeline_priority": "pipeline_priority",
    "effect_direction_consensus": "effect_direction",
}


def _merged_list_registry(config: dict[str, dict[str, object]]) -> dict[str, tuple[str, ...]]:
    """Resolve dropdown column headers to vocabulary tuples (supports alias keys)."""
    all_lists = phase0_excel_validation_lists()
    needed: set[str] = set()
    for sheet_cfg in config.values():
        for col in sheet_cfg.get("dropdowns", ()):
            needed.add(str(col))
    registry: dict[str, tuple[str, ...]] = {}
    for name in sorted(needed):
        key = _DROPDOWN_LIST_ALIASES.get(name, name)
        if key in all_lists:
            registry[name] = all_lists[key]
    return registry


def _resolved_clean_sheet_config(wb) -> dict[str, dict[str, object]]:
    """Merge literature sheets with Phase 0 sheets when present in workbook."""
    cfg = dict(CLEAN_SHEET_CONFIG)
    for name, sheet_cfg in PHASE0_VERIFICATION_SHEET_CONFIG.items():
        if name in wb.sheetnames:
            cfg[name] = sheet_cfg
    return cfg


def reorder_clean_workbook_sheets(wb) -> None:
    """Place Phase 0 sheets after Canonical_View; keep _FilterLists last."""
    ordered: list[Worksheet] = []
    seen: set[str] = set()
    for name in CLEAN_WORKBOOK_SHEET_ORDER:
        if name in wb.sheetnames:
            ordered.append(wb[name])
            seen.add(name)
    for name in wb.sheetnames:
        if name not in seen:
            ordered.append(wb[name])
    wb._sheets = ordered  # type: ignore[attr-defined]


def _clear_sheet_tables(ws: Worksheet) -> None:
    if not hasattr(ws, "tables"):
        return
    for name in list(ws.tables.keys()):
        del ws.tables[name]


def _clear_sheet_validations(ws: Worksheet) -> None:
    if ws.data_validations is None:
        return
    ws.data_validations.dataValidation = []


def _auto_column_widths(ws: Worksheet, *, max_width: int = MAX_COL_WIDTH) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = ws.cell(1, col_idx).value
        max_len = len(str(header or ""))
        for row_idx in range(2, min(ws.max_row + 1, 500)):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), max_width))
        ws.column_dimensions[letter].width = max(MIN_COL_WIDTH, min(max_len + 2, max_width))


def _freeze_header_row(ws: Worksheet) -> None:
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"


def _add_excel_table(ws: Worksheet, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    _clear_sheet_tables(ws)
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _ensure_filter_lists_sheet(wb, list_registry: dict[str, tuple[str, ...]]) -> dict[str, str]:
    """Write vocabularies to hidden sheet; return column header -> range formula."""
    if FILTER_LISTS_SHEET in wb.sheetnames:
        del wb[FILTER_LISTS_SHEET]
    ws = wb.create_sheet(FILTER_LISTS_SHEET)
    ws.sheet_state = "hidden"
    range_map: dict[str, str] = {}
    col_idx = 1
    for list_name, values in list_registry.items():
        if not values:
            continue
        ws.cell(1, col_idx, list_name)
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)
        letter = get_column_letter(col_idx)
        end_row = len(values) + 1
        range_map[list_name] = f"='{FILTER_LISTS_SHEET}'!${letter}$2:${letter}${end_row}"
        col_idx += 1
    return range_map


def _apply_dropdowns(
    ws: Worksheet,
    column_names: Sequence[str],
    header_map: dict[str, int],
    range_map: dict[str, str],
) -> None:
    _clear_sheet_validations(ws)
    if ws.max_row < 2:
        return
    for col_name in column_names:
        if col_name not in header_map:
            continue
        formula = range_map.get(col_name)
        if not formula:
            continue
        col_letter = get_column_letter(header_map[col_name])
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid filter value",
            error="Choose a value from the controlled vocabulary list.",
        )
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
        ws.add_data_validation(dv)


def _header_index(ws: Worksheet) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def _format_sheet(
    ws: Worksheet,
    *,
    table_name: str | None,
    dropdown_columns: Sequence[str],
    range_map: dict[str, str],
    apply_freeze: bool = True,
    apply_widths: bool = True,
) -> None:
    if apply_widths:
        _auto_column_widths(ws)
    if apply_freeze:
        _freeze_header_row(ws)
    if table_name:
        _add_excel_table(ws, table_name)
    if dropdown_columns:
        _apply_dropdowns(ws, dropdown_columns, _header_index(ws), range_map)


CLEAN_SHEET_CONFIG: dict[str, dict[str, object]] = {
    "0_Dashboard_Analysis_Control": {
        "table": "tbl_Dashboard",
        "dropdowns": (),
    },
    "1_Articles_Master": {
        "table": "tbl_ArticlesMaster",
        "dropdowns": (
            "study_filter_type",
            "ml_family",
            "pv_filter_label",
            "md_filter_label",
            "weights_filter",
        ),
    },
    "2_Main_Findings": {
        "table": "tbl_MainFindings",
        "dropdowns": ("study_filter_type", "target_domain", "target_dimension"),
    },
    "3_Confounders": {
        "table": "tbl_Confounders",
        "dropdowns": ("study_filter_type", "predictor_level", "predictor_category"),
    },
    "Canonical_View": {
        "table": "tbl_CanonicalView",
        "dropdowns": (
            "record_type",
            "study_filter_type",
            "canonical_method",
            "canonical_variable",
            "metadata_filter_flag",
            "effect_trend",
            "target_domain",
        ),
    },
}

ANALYTICAL_SHEET_CONFIG: dict[str, dict[str, object]] = {
    "Dashboard": {"table": "tbl_AnalyticalDashboard", "dropdowns": ()},
    "Study_Details": {
        "table": "tbl_StudyDetails",
        "dropdowns": ("Study_Filter_Type", "ML_Family"),
    },
    "Empirical_Findings": {
        "table": "tbl_EmpiricalFindings",
        "dropdowns": ("Study_Filter_Type", "ML_Family", "Target_Domain", "Target_Dimension"),
    },
    "Policy_Insights": {
        "table": "tbl_PolicyInsights",
        "dropdowns": (
            "Policy_Domain",
            "Policy_Impact_Direction",
            "Study_Filter_Type",
            "ML_Family",
        ),
    },
    "_Policy_Taxonomy_Map": {"table": "tbl_PolicyTaxonomy", "dropdowns": ()},
}


def _collect_needed_lists(config: dict[str, dict[str, object]]) -> dict[str, tuple[str, ...]]:
    all_lists = canonical_excel_validation_lists()
    needed: set[str] = set()
    for sheet_cfg in config.values():
        for col in sheet_cfg.get("dropdowns", ()):
            needed.add(str(col))
    return {name: all_lists[name] for name in sorted(needed) if name in all_lists}


def format_workbook(
    xlsx_path: Path,
    sheet_config: dict[str, dict[str, object]],
    *,
    only_sheets: Iterable[str] | None = None,
) -> None:
    """Apply Tables, dropdowns, freeze panes, and column widths."""
    path = xlsx_path.resolve()
    wb = load_workbook(path)
    range_map = _ensure_filter_lists_sheet(wb, _collect_needed_lists(sheet_config))
    targets = set(only_sheets) if only_sheets else set(sheet_config.keys())

    for sheet_name, cfg in sheet_config.items():
        if sheet_name not in wb.sheetnames or sheet_name not in targets:
            continue
        ws = wb[sheet_name]
        _format_sheet(
            ws,
            table_name=str(cfg["table"]) if cfg.get("table") else None,
            dropdown_columns=tuple(cfg.get("dropdowns", ())),  # type: ignore[arg-type]
            range_map=range_map,
        )

    wb.save(path)


def format_clean_workbook(xlsx_path: Path, *, only_sheets: Iterable[str] | None = None) -> None:
    """Format CLEAN workbook (Sheets 0–4, Canonical_View, and Phase 0 sheets if present)."""
    path = xlsx_path.resolve()
    wb = load_workbook(path)
    config = _resolved_clean_sheet_config(wb)
    range_map = _ensure_filter_lists_sheet(wb, _merged_list_registry(config))
    targets = set(only_sheets) if only_sheets else set(config.keys())
    for sheet_name, cfg in config.items():
        if sheet_name not in wb.sheetnames or sheet_name not in targets:
            continue
        ws = wb[sheet_name]
        _format_sheet(
            ws,
            table_name=str(cfg["table"]) if cfg.get("table") else None,
            dropdown_columns=tuple(cfg.get("dropdowns", ())),  # type: ignore[arg-type]
            range_map=range_map,
        )
    reorder_clean_workbook_sheets(wb)
    wb.save(path)


def format_phase0_verification_workbook(xlsx_path: Path) -> None:
    """Format only Sheets 5–7 and refresh _FilterLists."""
    format_workbook(xlsx_path, PHASE0_VERIFICATION_SHEET_CONFIG)


def format_analytical_workbook(xlsx_path: Path) -> None:
    """Format ILSA_Analytical_Meta_Analysis_Master.xlsx."""
    format_workbook(xlsx_path, ANALYTICAL_SHEET_CONFIG)
