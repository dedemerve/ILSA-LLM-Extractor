"""Tests for Excel workbook formatting helpers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.enrichment.excel_workbook_format import format_clean_workbook


def test_format_clean_workbook_adds_table_and_validation(tmp_path: Path) -> None:
    xlsx = tmp_path / "mini_clean.xlsx"
    df = pd.DataFrame(
        [
            {
                "file_name": "paper_a.pdf",
                "study_filter_type": "Empirical Study - Machine Learning",
                "ml_family": "Tree-Based / Ensemble Learning",
            }
        ]
    )
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        pd.DataFrame({"Section": ["Overview"], "Dimension": ["x"], "Value": [1], "Notes": [""]}).to_excel(
            writer, sheet_name="0_Dashboard_Analysis_Control", index=False
        )
        df.to_excel(writer, sheet_name="1_Articles_Master", index=False)
        pd.DataFrame({"file_name": ["paper_a.pdf"], "study_filter_type": ["Empirical Study - Machine Learning"], "target_domain": ["Mathematics"], "target_dimension": ["Cognitive Achievement"]}).to_excel(
            writer, sheet_name="2_Main_Findings", index=False
        )
        pd.DataFrame({"file_name": ["paper_a.pdf"], "study_filter_type": ["Empirical Study - Machine Learning"], "predictor_level": ["Student Level"], "predictor_category": ["Student: SES"]}).to_excel(
            writer, sheet_name="3_Confounders", index=False
        )

    format_clean_workbook(xlsx)

    wb = load_workbook(xlsx)
    assert "_FilterLists" in wb.sheetnames
    assert wb["_FilterLists"].sheet_state == "hidden"
    ws = wb["1_Articles_Master"]
    assert ws.freeze_panes == "A2"
    assert "tbl_ArticlesMaster" in ws.tables
    assert len(ws.data_validations.dataValidation) >= 1
    wb.close()
