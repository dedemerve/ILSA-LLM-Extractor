"""Build ILSA_Analytical_Meta_Analysis_Master.xlsx from the CLEAN dataset.

Policy-Oriented Extraction is deterministic (no API calls):
- canonical_variable (Canonical_View) is mapped to a fixed policy domain taxonomy.
- effect_trend distribution per (article, policy_domain) sets Policy_Impact_Direction.
- Country_Specific_Context unfolds countries_formatted to one row per country.
- Confidence_Score is computed from supporting-signal counts; missing signal -> [Not Explicitly Stated].

Output workbook has four sheets ready for Pivot/Slicer use:
  Dashboard, Study_Details, Empirical_Findings, Policy_Insights
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "outputs" / "ILSA_Meta_Analysis_Dataset_CLEAN.xlsx"
DST = ROOT / "outputs" / "ILSA_Analytical_Meta_Analysis_Master.xlsx"
TAXONOMY_JSON = ROOT / "outputs" / "policy_domain_taxonomy.json"

NOT_STATED = "[Not Explicitly Stated]"
INTERNATIONAL = "International / Cross-Country"

# canonical_variable -> policy domain
POLICY_DOMAIN_MAP: dict[str, str] = {
    "Curriculum_Instruction": "Curriculum & Instruction",
    "Homework_Study_Time": "Curriculum & Instruction",
    "Assessment_Methodology": "Curriculum & Instruction",
    "Assessment_Practices": "Curriculum & Instruction",
    "Metacognition_Strategies": "Curriculum & Instruction",
    "Teacher_Quality_Practices": "Teacher Policy & Workforce",
    "Teacher_Efficacy_Workforce": "Teacher Policy & Workforce",
    "School_Resources": "School Finance & Resources",
    "School_Efficiency_Climate": "School Finance & Resources",
    "School_Type_Composition": "School Finance & Resources",
    "SES": "Socio-Economic Status & Equity",
    "Academic_Resilience": "Socio-Economic Status & Equity",
    "Labor_Market_Outcomes": "Socio-Economic Status & Equity",
    "ICT_Usage": "Digital Learning & ICT",
    "Process_Data_Log": "Digital Learning & ICT",
    "Immigration_Language": "Language & Immigrant Policy",
    "Language_Proficiency": "Language & Immigrant Policy",
    "Parental_Involvement": "Family Engagement",
    "Wellbeing_Psychological": "Student Wellbeing & Climate",
    "Wellbeing_Mental_Health": "Student Wellbeing & Climate",
    "Anxiety_Stress": "Student Wellbeing & Climate",
    "Belonging_School_Climate": "Student Wellbeing & Climate",
    "Student_Motivation": "Motivation & Self-Efficacy",
    "Self_Efficacy": "Motivation & Self-Efficacy",
    "Gender_Demographics": "Gender & Demographic Equity",
    "System_Policy": "System Governance",
    "Peer_Effects": "Peer Effects & Tracking",
    "Dropout_Retention": "Retention & Pathways",
    "Occupational_Expectation": "Retention & Pathways",
    "Civic_Achievement": "Civic Education",
    "Civic_Engagement": "Civic Education",
}
# Variables we treat as outcomes / non-policy levers (excluded from Policy_Insights).
OUTCOME_VARIABLES: set[str] = {
    "Math_Achievement",
    "Reading_Achievement",
    "Science_Achievement",
    "Problem_Solving_Achievement",
    "Prior_Achievement",
    "Theoretical_and_Meta_Synthesis",
    "Psychometrics_Process",
    "Uncategorized_Contextual",
}

POLICY_KEYWORDS = re.compile(
    r"\b(polic(?:y|ies)|reform|curricul(?:um|a)|intervention|recommend|implement|"
    r"invest|funding|resource allocation|equity|targeted|professional development|"
    r"teacher training|early childhood|tracking|streaming|inclusive|"
    r"language support|integration|digital literacy|ict integration|"
    r"assessment design|standards?|accountability)\b",
    re.IGNORECASE,
)

NEGATIVE_DIRECTION_WORDS = re.compile(
    r"\b(reduc(?:e|ed|ing)|mitigat(?:e|ed|ing)|narrow|lower|prevent|address|"
    r"corrective|remediate|gap|disadvantage|disparit(?:y|ies))\b",
    re.IGNORECASE,
)


def _norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def _split_countries(field: str) -> list[str]:
    if not field:
        return []
    txt = field.replace(";", ",").replace("|", ",")
    parts = [p.strip() for p in txt.split(",") if p.strip()]
    if not parts:
        return []
    # If looks like a free-text summary rather than a clean list, keep one item.
    if len(parts) == 1 and len(parts[0]) > 60:
        return [parts[0]]
    return parts


def _ml_family(value: str) -> str:
    return value if value else "Not Reported"


def _direction_from_trends(trends: Counter[str]) -> str:
    pos = trends.get("Positive", 0)
    neg = trends.get("Negative", 0)
    null = trends.get("Null", 0)
    if pos == 0 and neg == 0:
        return "Descriptive / Null" if null else "Not Reported"
    if pos > 0 and neg == 0:
        return "Improving (positive predictor)"
    if neg > 0 and pos == 0:
        return "Corrective (negative association)"
    return "Mixed / Context-dependent"


def _confidence(
    n_confounder_signals: int,
    n_finding_signals: int,
    has_nonnull_trend: bool,
    has_policy_language: bool,
) -> float:
    score = 0.0
    if n_confounder_signals:
        score += min(0.30 + 0.05 * (n_confounder_signals - 1), 0.45)
    if n_finding_signals:
        score += min(0.15 + 0.05 * (n_finding_signals - 1), 0.30)
    if has_nonnull_trend:
        score += 0.15
    if has_policy_language:
        score += 0.10
    return round(min(score, 1.0), 2)


def _action_phrase(domain: str, direction: str) -> str:
    base = {
        "Curriculum & Instruction": "curriculum and instructional practice",
        "Teacher Policy & Workforce": "teacher quality and workforce policy",
        "School Finance & Resources": "school resource allocation",
        "Socio-Economic Status & Equity": "SES-targeted equity policy",
        "Digital Learning & ICT": "ICT and digital learning policy",
        "Language & Immigrant Policy": "language support and immigrant integration",
        "Family Engagement": "family engagement policy",
        "Student Wellbeing & Climate": "student wellbeing and school climate",
        "Motivation & Self-Efficacy": "motivation and self-efficacy programs",
        "Gender & Demographic Equity": "gender and demographic equity",
        "System Governance": "system-level governance",
        "Peer Effects & Tracking": "peer composition and tracking policy",
        "Retention & Pathways": "retention and post-school pathways",
        "Civic Education": "civic education policy",
    }.get(domain, domain.lower())
    if direction.startswith("Improving"):
        return f"Strengthen / scale {base}"
    if direction.startswith("Corrective"):
        return f"Address adverse effects in {base}"
    if direction.startswith("Mixed"):
        return f"Calibrate {base} (context-sensitive)"
    if direction.startswith("Descriptive"):
        return f"Monitor {base}"
    return NOT_STATED


def load_source() -> dict[str, pd.DataFrame]:
    sheets = pd.read_excel(SRC, sheet_name=None, engine="openpyxl")
    print("Source sheets:")
    for name, df in sheets.items():
        print(f"  {name}: {len(df):>5d} rows  x  {df.shape[1]:>3d} cols")
    return sheets


def build_taxonomy(canonical: pd.DataFrame) -> pd.DataFrame:
    cv = canonical[canonical["canonical_variable"].notna()].copy()
    cv = cv[cv["canonical_variable"] != "[IGNORE]"]
    counts = cv["canonical_variable"].value_counts().to_dict()
    rows = []
    for var, n in counts.items():
        rows.append(
            {
                "canonical_variable": var,
                "occurrences": n,
                "policy_domain": POLICY_DOMAIN_MAP.get(var, NOT_STATED if var in OUTCOME_VARIABLES else NOT_STATED),
                "treated_as_outcome": var in OUTCOME_VARIABLES,
            }
        )
    return pd.DataFrame(rows)


def aggregate_per_article(sheets: dict[str, pd.DataFrame]) -> dict[str, dict]:
    articles = sheets["1_Articles_Master"].copy()
    findings = sheets["2_Main_Findings"].copy()
    canonical = sheets["Canonical_View"].copy()

    art_idx = articles.set_index("file_name", drop=False)
    per_article: dict[str, dict] = {}

    for file_name, row in art_idx.iterrows():
        per_article[file_name] = {
            "row": row,
            "countries": _split_countries(_norm(row.get("countries_formatted"))) or [INTERNATIONAL],
            "predictor_vars": Counter(),
            "finding_trends": defaultdict(Counter),
            "policy_text_blob": " ".join(
                filter(
                    None,
                    [
                        _norm(row.get("primary_finding")),
                        _norm(row.get("outcome_summary")),
                        _norm(row.get("title")),
                    ],
                )
            ),
        }

    for _, r in canonical.iterrows():
        fname = _norm(r.get("file_name"))
        var = _norm(r.get("canonical_variable"))
        trend = _norm(r.get("effect_trend"))
        if not fname or fname not in per_article:
            continue
        if not var or var == "[IGNORE]":
            continue
        if r["record_type"] == "confounder":
            per_article[fname]["predictor_vars"][var] += 1
        elif r["record_type"] == "finding":
            per_article[fname]["finding_trends"][var][trend or "Null"] += 1

    return per_article


def build_study_details(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    src = sheets["1_Articles_Master"]
    cols = [
        ("file_name", "File"),
        ("doi", "DOI"),
        ("title", "Title"),
        ("authors", "Authors"),
        ("year", "Year"),
        ("venue", "Venue"),
        ("publication_type", "Publication_Type"),
        ("source_category", "Source_Category"),
        ("corpus_source", "Corpus_Source"),
        ("research_design_type", "Research_Design"),
        ("ml_primary", "ML_Primary"),
        ("ml_family", "ML_Family"),
        ("ml_all_techniques", "ML_All_Techniques"),
        ("sample_size", "Sample_Size"),
        ("total_students", "Total_Students"),
        ("countries_formatted", "Countries"),
        ("study_filter_type", "Study_Filter_Type"),
        ("document_class", "Document_Class"),
        ("pv_filter_label", "Plausible_Values_Handling"),
        ("md_filter_label", "Missing_Data_Handling"),
        ("weights_filter", "Weights_Handling"),
        ("outcome_summary", "Outcome_Summary"),
        ("primary_finding", "Primary_Finding"),
    ]
    df = src[[c for c, _ in cols]].copy()
    df.columns = [label for _, label in cols]
    df["ML_Family"] = df["ML_Family"].fillna("").map(_ml_family)
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")
    return df


def build_empirical_findings(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    findings = sheets["2_Main_Findings"].copy()
    articles = sheets["1_Articles_Master"][
        ["file_name", "ml_primary", "ml_family", "countries_formatted", "year"]
    ]
    canonical = sheets["Canonical_View"]

    # Bring canonical variable + trend from findings rows of Canonical_View.
    cv_find = canonical[canonical["record_type"] == "finding"][
        ["file_name", "raw_label", "canonical_variable", "effect_trend"]
    ].rename(
        columns={
            "raw_label": "target_or_predictor_label",
            "canonical_variable": "Canonical_Variable",
            "effect_trend": "Effect_Trend",
        }
    )

    df = findings.merge(articles, on="file_name", how="left")
    df = df.rename(
        columns={
            "file_name": "File",
            "doi": "DOI",
            "dataset_used": "Dataset_Used",
            "target_variable": "Target_Variable",
            "target_domain": "Target_Domain",
            "target_dimension": "Target_Dimension",
            "top_predictors": "Top_Predictors",
            "performance_metrics": "Performance_Metrics",
            "standardized_conclusion": "Standardized_Conclusion",
            "effect_size": "Effect_Size",
            "primary_finding": "Primary_Finding",
            "predictor_filter_categories": "Predictor_Filter_Categories",
            "publication_type": "Publication_Type",
            "source_category": "Source_Category",
            "study_filter_type": "Study_Filter_Type",
            "document_class": "Document_Class",
            "ml_primary": "ML_Primary",
            "ml_family": "ML_Family",
            "countries_formatted": "Countries",
            "year": "Year",
        }
    )
    df["ML_Family"] = df["ML_Family"].fillna("").map(_ml_family)
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")

    # Per-file Canonical_Variable summary (positive/negative/null counts).
    cv_summary = (
        cv_find.groupby(["file_name", "Canonical_Variable", "Effect_Trend"]).size().unstack(fill_value=0).reset_index()
        if not cv_find.empty
        else pd.DataFrame()
    )
    out_cols = [
        "File", "DOI", "Year", "Countries", "Dataset_Used",
        "Target_Domain", "Target_Dimension", "Target_Variable",
        "Top_Predictors", "Predictor_Filter_Categories",
        "Effect_Size", "Performance_Metrics", "Standardized_Conclusion",
        "ML_Primary", "ML_Family", "Study_Filter_Type", "Document_Class",
        "Publication_Type", "Source_Category", "Primary_Finding",
    ]
    return df[out_cols]


def build_policy_insights(
    sheets: dict[str, pd.DataFrame],
    per_article: dict[str, dict],
) -> pd.DataFrame:
    articles = sheets["1_Articles_Master"]
    rows: list[dict] = []

    for _, art in articles.iterrows():
        file_name = art["file_name"]
        agg = per_article.get(file_name)
        if agg is None:
            continue

        countries = agg["countries"]
        text_blob = agg["policy_text_blob"]
        has_policy_language = bool(POLICY_KEYWORDS.search(text_blob))

        domain_signals: dict[str, dict] = defaultdict(
            lambda: {"confounder_vars": [], "finding_trends": Counter()}
        )

        for var, n in agg["predictor_vars"].items():
            if var in OUTCOME_VARIABLES:
                continue
            dom = POLICY_DOMAIN_MAP.get(var)
            if not dom:
                continue
            domain_signals[dom]["confounder_vars"].append(var)

        for var, trends in agg["finding_trends"].items():
            if var in OUTCOME_VARIABLES:
                continue
            dom = POLICY_DOMAIN_MAP.get(var)
            if not dom:
                continue
            for t, n in trends.items():
                domain_signals[dom]["finding_trends"][t] += n

        # No policy signal at all -> still emit a sentinel row so the article appears.
        if not domain_signals:
            for country in countries:
                rows.append(
                    {
                        "File": file_name,
                        "DOI": _norm(art.get("doi")),
                        "Year": int(art["year"]) if pd.notna(art.get("year")) else None,
                        "Country": country,
                        "Policy_Domain": NOT_STATED,
                        "Policy_Impact_Direction": NOT_STATED,
                        "Recommended_Action": NOT_STATED,
                        "Supporting_Variables": NOT_STATED,
                        "Trend_Counts_Pos_Neg_Null": "0/0/0",
                        "Confidence_Score": 0.0,
                        "Country_Specific_Context": NOT_STATED,
                        "ML_Primary": _norm(art.get("ml_primary")),
                        "ML_Family": _ml_family(_norm(art.get("ml_family"))),
                        "Study_Filter_Type": _norm(art.get("study_filter_type")),
                        "Publication_Type": _norm(art.get("publication_type")),
                        "Source_Category": _norm(art.get("source_category")),
                    }
                )
            continue

        for domain, sig in domain_signals.items():
            trends = sig["finding_trends"]
            direction = _direction_from_trends(trends)
            confidence = _confidence(
                n_confounder_signals=len(set(sig["confounder_vars"])),
                n_finding_signals=sum(trends.values()),
                has_nonnull_trend=(trends.get("Positive", 0) + trends.get("Negative", 0)) > 0,
                has_policy_language=has_policy_language,
            )
            action = _action_phrase(domain, direction)
            supp_vars = ", ".join(sorted(set(sig["confounder_vars"]))) or NOT_STATED
            trend_str = f"{trends.get('Positive', 0)}/{trends.get('Negative', 0)}/{trends.get('Null', 0)}"

            for country in countries:
                ctx = (
                    f"{country}: {action.lower()}"
                    if country != INTERNATIONAL
                    else f"Cross-country: {action.lower()}"
                )
                rows.append(
                    {
                        "File": file_name,
                        "DOI": _norm(art.get("doi")),
                        "Year": int(art["year"]) if pd.notna(art.get("year")) else None,
                        "Country": country,
                        "Policy_Domain": domain,
                        "Policy_Impact_Direction": direction,
                        "Recommended_Action": action,
                        "Supporting_Variables": supp_vars,
                        "Trend_Counts_Pos_Neg_Null": trend_str,
                        "Confidence_Score": confidence,
                        "Country_Specific_Context": ctx,
                        "ML_Primary": _norm(art.get("ml_primary")),
                        "ML_Family": _ml_family(_norm(art.get("ml_family"))),
                        "Study_Filter_Type": _norm(art.get("study_filter_type")),
                        "Publication_Type": _norm(art.get("publication_type")),
                        "Source_Category": _norm(art.get("source_category")),
                    }
                )

    return pd.DataFrame(rows)


def build_dashboard(
    study: pd.DataFrame,
    findings: pd.DataFrame,
    policy: pd.DataFrame,
) -> pd.DataFrame:
    sections: list[tuple[str, str, str]] = []

    def add(section: str, dim: str, value):
        sections.append((section, dim, str(value)))

    add("Corpus", "Total_Articles", len(study))
    add("Corpus", "Total_Empirical_Findings", len(findings))
    add("Corpus", "Total_Policy_Insight_Rows", len(policy))
    add("Corpus", "Years_Covered", f"{int(study['Year'].min())} - {int(study['Year'].max())}")
    add("Corpus", "Unique_Countries", policy["Country"].nunique())

    for label, value in policy["Policy_Domain"].value_counts().head(15).items():
        add("Policy_Domain", label, value)
    for label, value in policy["Policy_Impact_Direction"].value_counts().items():
        add("Policy_Impact_Direction", label, value)
    for label, value in study["ML_Family"].value_counts().head(15).items():
        add("ML_Family", label, value)
    for label, value in study["Source_Category"].value_counts().head(10).items():
        add("Source_Category", label, value)
    for label, value in policy["Country"].value_counts().head(20).items():
        add("Top_Countries", label, value)
    for label, value in policy["Confidence_Score"].round(1).value_counts().sort_index().items():
        add("Confidence_Distribution", str(label), value)

    return pd.DataFrame(sections, columns=["Section", "Dimension", "Value"])


def _write_workbook(
    dashboard: pd.DataFrame,
    study: pd.DataFrame,
    findings: pd.DataFrame,
    policy: pd.DataFrame,
    taxonomy: pd.DataFrame,
) -> None:
    DST.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(DST, engine="openpyxl") as xw:
        dashboard.to_excel(xw, sheet_name="Dashboard", index=False)
        study.to_excel(xw, sheet_name="Study_Details", index=False)
        findings.to_excel(xw, sheet_name="Empirical_Findings", index=False)
        policy.to_excel(xw, sheet_name="Policy_Insights", index=False)
        taxonomy.to_excel(xw, sheet_name="_Policy_Taxonomy_Map", index=False)

    # Add Excel Tables (so Pivot/Slicer "just works")
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = load_workbook(DST)
    for sheet_name, table_name in (
        ("Study_Details", "tbl_Study"),
        ("Empirical_Findings", "tbl_Findings"),
        ("Policy_Insights", "tbl_Policy"),
        ("_Policy_Taxonomy_Map", "tbl_Taxonomy"),
    ):
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 1:
            continue
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(t)
    wb.save(DST)


def main() -> None:
    sheets = load_source()
    per_article = aggregate_per_article(sheets)

    taxonomy = build_taxonomy(sheets["Canonical_View"])
    TAXONOMY_JSON.write_text(json.dumps(POLICY_DOMAIN_MAP, indent=2, ensure_ascii=False))
    print(f"\nTaxonomy: {taxonomy['policy_domain'].nunique()} unique domains, "
          f"{len(taxonomy)} canonical variables mapped (incl. outcome-only).")

    study = build_study_details(sheets)
    findings = build_empirical_findings(sheets)
    policy = build_policy_insights(sheets, per_article)
    dashboard = build_dashboard(study, findings, policy)

    print("\nFinal row counts:")
    print(f"  Dashboard:          {len(dashboard):>5d}")
    print(f"  Study_Details:      {len(study):>5d}")
    print(f"  Empirical_Findings: {len(findings):>5d}")
    print(f"  Policy_Insights:    {len(policy):>5d}")

    pct_stated = (policy["Policy_Domain"] != NOT_STATED).mean()
    print(f"\nPolicy coverage: {pct_stated:.1%} of policy rows have a stated domain.")
    print("Confidence distribution:")
    print(policy["Confidence_Score"].describe().round(3).to_string())

    _write_workbook(dashboard, study, findings, policy, taxonomy)
    print(f"\nWrote {DST.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
